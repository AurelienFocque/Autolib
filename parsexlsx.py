from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import requests
from time import *
f="location_recap.xlsx"
wb = load_workbook(filename = f)
ws=wb["Sheet"]
def montant():
  ret=0.0
  for col in ws.get_squared_range(9, 2, 9, ws.max_row):
    for cell in col:
      ret+=float(cell.value)
  return ret
def temps():
  ret=0.0
  for col in ws.get_squared_range(4, 2, 4, ws.max_row):
    for cell in col:
      ret+=float(cell.value)
  return ret
#need to provide google map api key for this function to work
API_KEY="0"
def distance():
  ws["J1"].value="Distance en metres API GOOGLE"
  distance=0
  if(API_KEY=="0"):
    return 0
  url="https://maps.googleapis.com/maps/api/distancematrix/json?origins=&destinations=&language=fr-FR&key="+API_KEY
  s=requests.Session()
  departs=[]
  for col in ws.get_squared_range(6, 2, 6, ws.max_row):
    for cell in col:
      departs.append(cell.value)
  arrivees=[]
  for col in ws.get_squared_range(7, 2, 7, ws.max_row):
    for cell in col:
      arrivees.append(cell.value)
  for k in range(0,len(arrivees)):
    r=s.get(url.replace("origins=","origins="+departs[k]).replace("destinations=","destinations="+arrivees[k]))
    dtext=r.text.split('m",\n                  "value" : ')[1].split("\n")[0]
    print(dtext)
    ws["J"+str(k+2)].value=int(dtext)
    distance+=int(dtext)
  return (distance)
print("Nombre de jours : ")
t=temps()
print(t)
print("Prix total : ")
m=montant()
print(m)
print("kilometres parcourus : ")
d=distance()
print(d/1000.0)
ws["C"+str(ws.max_row+1)].value=("%.2f" %(t*24))+" Heures"
ws["I"+str(ws.max_row)].value=str(m)+" Euros"
ws["J"+str(ws.max_row)].value=str(d/1000.0)+" Km"
date = datetime.datetime.now()
wb.save("recap_complet"+str(date)+".xlsx")
